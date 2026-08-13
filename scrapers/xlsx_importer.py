"""二手房成交 xlsx 导入器（需在本机运行：pip install -r requirements.txt）。

为什么单独成脚本：链家等平台的「城市成交记录」多为 Excel(xlsx) 而非 CSV，
且解析需要 openpyxl / pandas。沙箱环境无法安装这些依赖，因此本脚本设计为由
你在本地执行，把更多城市的真实二手房房源级数据补充进 SQLite。

已验证可用的数据源（GitHub，可按需 clone 后把 xlsx 放进 data/raw/）：
  - sczhengyabin/Lianjia_House_Info  -> 成都_成交记录——20170830.xlsx
    （字段：小区名称/行政区域/户类型/建筑面积/成交价/挂牌价/单价/
           成交日期/朝向/装修/楼层/总楼层/建筑年份）

用法：
    cd RealEstateAI
    pip install -r requirements.txt        # 含 openpyxl
    python scrapers/xlsx_importer.py               # 导入全部 SOURCES
    python scrapers/xlsx_importer.py --clear      # 清旧后重导（幂等）

新增城市：在下方 SOURCES 增加一项并填写列名映射即可（COL_MAP 见各城市表头）。
"""
import os
import re
import sys

# 复用 dataset_importer 里的归一化工具，保证与既有数据口径一致
from dataset_importer import parse_floor, normalize_layout, to_float

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
DB_PATH = os.path.join(_ROOT, "data", "realestate.db")
RAW_DIR = os.path.join(_ROOT, "data", "raw")

# 成都成交记录（sczhengyabin/Lianjia_House_Info）的列名映射
# xlsx 表头 -> House 字段
CHENGDU_COLS = {
    "city": "成都",
    "community": "小区名称",   # C
    "region": "行政区域",      # D
    "rooms": "户类型",         # F
    "area": "建筑面积",        # G
    "price": "成交价",         # H  (单位：万元)
    "list_price": "挂牌价",    # I  (单位：万元)
    "unit_price": "单价",      # J  (单位：元/㎡)
    "deal_date": "成交日期",   # L  (格式 2017.08.07)
    "orientation": "朝向",     # N
    "decoration": "装修",      # O
    "floor": "楼层",           # Q  (高楼层/中楼层/低楼层)
    "floor_total": "总楼层",   # R
    "building_year": "建筑年份",  # S
}

# 价格单位换算：成交价/挂牌价按“万元”入库；如你的 xlsx 单位是“元”，把这里改成 10000
PRICE_DIV = 1  # 成交价列已是「万元」

SOURCES = [
    {
        "file": "chengdu_chengjiao.xlsx",
        "city": "成都",
        "cols": CHENGDU_COLS,
    },
]


def _parse_deal_year(s):
    m = re.search(r"(\d{4})", str(s or ""))
    return int(m.group(1)) if m else None


def import_source(src, clear=False):
    import sqlite3
    from openpyxl import load_workbook

    path = os.path.join(RAW_DIR, src["file"])
    if not os.path.exists(path):
        print(f"⚠️  跳过缺失文件：{path}")
        return 0, 0

    cols = src["cols"]
    city = cols["city"]
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """CREATE TABLE IF NOT EXISTS houses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            city TEXT, region TEXT, community TEXT, year INTEGER,
            title TEXT, price REAL, unit_price REAL, area REAL,
            rooms TEXT, floor_info TEXT, orientation TEXT,
            decoration TEXT, building_year INTEGER,
            property_type TEXT DEFAULT '二手房',
            description TEXT, url TEXT UNIQUE, deal_id TEXT, crawled_at TEXT, created_at TEXT
        )"""
    )
    if clear:
        conn.execute("DELETE FROM houses WHERE url LIKE ?", (f"dataset://{city}-xlsx/%",))
        conn.commit()

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"⚠️  {src['file']} 无数据")
        return 0, 0
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    name_to_idx = {name: i for i, name in enumerate(header)}

    def col(name):
        return name_to_idx.get(cols.get(name, ""))

    inserted = 0
    skipped = 0
    batch = []
    now = __import__("datetime").datetime.now().isoformat()
    idx_area, idx_price, idx_unit = col("area"), col("price"), col("unit_price")
    idx_comm, idx_region = col("community"), col("region")
    idx_rooms, idx_year = col("rooms"), col("deal_date")
    idx_orient, idx_deco = col("orientation"), col("decoration")
    idx_floor, idx_byear = col("floor"), col("building_year")

    for r_i, row in enumerate(rows[1:], start=0):
        community = str(row[idx_comm]).strip() if idx_comm is not None and row[idx_comm] is not None else ""
        region = str(row[idx_region]).strip() if idx_region is not None and row[idx_region] is not None else ""
        area = to_float(row[idx_area]) if idx_area is not None else None
        price = to_float(row[idx_price]) if idx_price is not None else None
        unit = to_float(row[idx_unit]) if idx_unit is not None else None
        if price is None or area is None or area <= 0:
            skipped += 1
            continue
        price = price / PRICE_DIV if PRICE_DIV else price
        rooms = normalize_layout(str(row[idx_rooms]).strip()) if idx_rooms is not None and row[idx_rooms] else None
        year = _parse_deal_year(row[idx_year]) if idx_year is not None and row[idx_year] else None
        orientation = str(row[idx_orient]).strip() if idx_orient is not None and row[idx_orient] else None
        decoration = str(row[idx_deco]).strip() if idx_deco is not None and row[idx_deco] else None
        floor_info = parse_floor(str(row[idx_floor]).strip()) if idx_floor is not None and row[idx_floor] else None
        building_year = int(to_float(row[idx_byear])) if idx_byear is not None and row[idx_byear] else None
        title = f"{community} {rooms or ''}".strip()
        desc = f"{title}，位于{region}，{rooms or ''}，{orientation or ''}朝向，{floor_info or ''}，建筑面积{area:.1f}平米，成交价{price}万，单价约{int(unit) if unit else 0}元/平。"
        url = f"dataset://{city}-xlsx/{r_i}"
        batch.append((
            city, region, community, year, title, price, unit, area, rooms,
            floor_info, orientation, decoration, building_year,
            "二手房", desc, url, now, now,
        ))
        if len(batch) >= 2000:
            conn.executemany(
                """INSERT OR IGNORE INTO houses
                   (city, region, community, year, title, price, unit_price, area, rooms,
                    floor_info, orientation, decoration, building_year,
                    property_type, description, url, crawled_at, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                batch,
            )
            inserted += len(batch)
            batch.clear()

    if batch:
        conn.executemany(
            """INSERT OR IGNORE INTO houses
               (city, region, community, year, title, price, unit_price, area, rooms,
                floor_info, orientation, decoration, building_year,
                property_type, description, url, crawled_at, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            batch,
        )
        inserted += len(batch)
    conn.commit()
    conn.close()
    print(f"  ✅ {src['file']} ({city}): 新增 {inserted} 条, 跳过 {skipped} 条")
    return inserted, skipped


def main():
    clear = "--clear" in sys.argv
    total_ins = 0
    total_skip = 0
    for src in SOURCES:
        ins, skip = import_source(src, clear=clear)
        total_ins += ins
        total_skip += skip
    print(f"🎉 xlsx 导入完成: 共新增 {total_ins} 条真实二手房数据, 跳过 {total_skip} 条")


if __name__ == "__main__":
    main()
